# Copyright (c) 2023 42dot. All rights reserved.
import argparse
import os.path

import torch
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False
torch.backends.cuda.matmul.allow_tf32 = False
    
import utils
from models import VFDepthAlgo
from trainer import VFDepthTrainer


def parse_args():
    parser = argparse.ArgumentParser(description='VFdepth evaluation script')
    parser.add_argument('--config_file', default ='./configs/surround_fusion.yaml', type=str, help='Config yaml file')
    parser.add_argument('--weight_path', default = None, type=str, help='Pretrained weight path')
    parser.add_argument('--nusc_type', default = None, type=str, help='Pretrained weight path')
    parser.add_argument('--port', default = None, type=str, help='Pretrained weight path')
    parser.add_argument('--post_process', action='store_true')
    parser.add_argument('--overlap', action='store_true')
    args = parser.parse_args()
    return args


def v_test_1(cfg):
    print("Evaluating")
    model = VFDepthAlgo(cfg, 0)
    trainer = VFDepthTrainer(cfg, 0, use_tb = False)
    trainer.evaluate(model, vis_results = cfg['eval']['eval_visualize'])

    
def v_test_ddp(rank, cfg,args):
    print("Evaluating")
    utils.setup_ddp(rank, cfg['ddp']['world_size'],cfg)

    import openpyxl

    # 获取 工作簿对象

    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    worksheet.title = "metric_metrics"
    worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
    worksheet2.title = "median_metrics"

    worksheets = []
    for i,cam in enumerate(cfg['data']['cameras']):
        worksheets.append(workbook.create_sheet())
        worksheets[i].title=cam
    for i,metric in enumerate(['epoch','abs_rel', 'sq_rel', 'rms', 'log_rms', 'a1', 'a2', 'a3']):
        worksheet.cell(1,i+1,metric)
        worksheet2.cell(1,i+1,metric)
        for j, cam in enumerate(cfg['data']['cameras']):
            worksheets[j].cell(1,i+1,metric)

    start = 1
    end = 20
    if args.weight_path is not None:
        start =int(args.weight_path.split('_')[-1])
        end =int(args.weight_path.split('_')[-1])+1
    for ep in range(start,end):
        model = VFDepthAlgo(cfg, rank)
        model.load_weights_dir = model.load_weights_dir[:-2]+str(ep)
        print(model.load_weights_dir )
        if os.path.isdir(model.load_weights_dir):
            trainer = VFDepthTrainer(cfg, rank, use_tb=False)
            avg_depth_eval_metric,avg_depth_eval_median,avg_depth_eval_metric_cams = \
                            trainer.evaluate(model,vis_results = cfg['eval']['eval_visualize'])
            for i, metric in enumerate(['epoch', 'abs_rel', 'sq_rel', 'rms', 'log_rms', 'a1', 'a2', 'a3']):
                if i==0:
                    worksheet.cell(ep + 2, i + 1, ep)
                    worksheet2.cell(ep + 2 , i + 1, ep)
                    for j, cam in enumerate(cfg['data']['cameras']):
                        worksheets[j].cell(ep + 2, i + 1, ep)
                else:
                    worksheet.cell(ep+2, i + 1, avg_depth_eval_metric[metric])
                    worksheet2.cell(ep+2, i + 1, avg_depth_eval_median[metric])
                    for j, cam in enumerate(cfg['data']['cameras']):
                        worksheets[j].cell(ep+2, i + 1, avg_depth_eval_metric_cams[j][metric])


    if cfg['eval'].get('type') is not None:
        tt = '_'+cfg['eval'].get('type')
    else:
        tt=""
    if cfg['eval'].get('overlap') is True:
        tt+='_overlap'
    if cfg['eval'].get('post_process'):
        workbook.save(filename=os.path.join(model.log_path, 'results_pp_my_depth{}.xlsx'.format(tt)))
    else:
        workbook.save(filename=os.path.join(model.log_path, 'results_my_depth{}.xlsx'.format(tt)))
    utils.clear_ddp()
    

if __name__ == '__main__':
    args = parse_args()
    cfg = utils.get_config(args.config_file, mode='eval', weight_path = args.weight_path)
    if args.post_process:
        cfg['eval']['post_process'] = True
    if args.overlap:
        cfg['eval']['overlap'] = True
    if args.nusc_type:
        cfg['eval']['type'] = args.nusc_type
    if args.port:
        cfg['ddp']['port'] = args.port
    # Evaluating on DDP trained model
    if cfg['ddp']['ddp_enable'] == True:
        import torch.multiprocessing as mp
        mp.spawn(v_test_ddp, nprocs=cfg['ddp']['world_size'], args=(cfg,args,),join=True)
    else:
        v_test_1(cfg)
